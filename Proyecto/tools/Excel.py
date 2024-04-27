import sys
from PyQt5.QtWidgets import QFileDialog
from openpyxl import Workbook
from tools.tools import Exito

def export_to_excel(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Datos"

        # Agregar títulos de columna
        for col in range(self.TablaVales.columnCount()):
            title = self.TablaVales.horizontalHeaderItem(col).text()
            sheet.cell(row=1, column=col + 2).value = title

        # Agregar datos de la tabla
        for row in range(self.TablaVales.rowCount()):
            sheet.cell(row=row + 2, column=1).value = row + 1  # Enumeración
            for col in range(self.TablaVales.columnCount()):
                item = self.TablaVales.item(row, col)
                if item is not None:
                    sheet.cell(row=row + 2, column=col + 2).value = item.text()

        # Ajustar el ancho de las columnas para que quepan los datos
        for col in range(self.TablaVales.columnCount() + 1):
            max_length = 0
            for row in range(self.TablaVales.rowCount() + 1):
                if sheet.cell(row=row + 1, column=col + 1).value:
                    max_length = max(max_length, len(str(sheet.cell(row=row + 1, column=col + 1).value)))
            sheet.column_dimensions[chr(65 + col)].width = max_length + 2

        # Guardar el archivo de Excel
        filename, _ = QFileDialog.getSaveFileName(self, "Guardar archivo", "", "Excel Files (*.xlsx)")
        if filename:
            workbook.save(filename)
            Exito("Archivo guardado exitosamente.")



def export_to_excel1(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Datos"

        # Agregar títulos de columna
        for col in range(self.TablaGastos.columnCount()):
            title = self.TablaGastos.horizontalHeaderItem(col).text()
            sheet.cell(row=1, column=col + 2).value = title

        # Agregar datos de la tabla
        for row in range(self.TablaGastos.rowCount()):
            sheet.cell(row=row + 2, column=1).value = row + 1  # Enumeración
            for col in range(self.TablaGastos.columnCount()):
                item = self.TablaGastos.item(row, col)
                if item is not None:
                    sheet.cell(row=row + 2, column=col + 2).value = item.text()

        # Ajustar el ancho de las columnas para que quepan los datos
        for col in range(self.TablaGastos.columnCount() + 1):
            max_length = 0
            for row in range(self.TablaGastos.rowCount() + 1):
                if sheet.cell(row=row + 1, column=col + 1).value:
                    max_length = max(max_length, len(str(sheet.cell(row=row + 1, column=col + 1).value)))
            sheet.column_dimensions[chr(65 + col)].width = max_length + 2

        # Guardar el archivo de Excel
        filename, _ = QFileDialog.getSaveFileName(self, "Guardar archivo", "", "Excel Files (*.xlsx)")
        if filename:
            workbook.save(filename)
            Exito("Archivo guardado exitosamente.")